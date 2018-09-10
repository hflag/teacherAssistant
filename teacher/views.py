from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.mixins import LoginRequiredMixin
from django.views.generic.base import View
from django.http import JsonResponse, HttpResponse
import time
import random
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

from .models import Schedule, Student, Act, StudentActScore, Banji
from .forms import CustomerForm


class IndexView(View):
    def get(self, request):
        return render(request, 'teacher/index.html')


class ClassActive(LoginRequiredMixin, View):
    def post(self, request):
        act_name = request.POST.get('act_name')
        banji_name = request.POST.get('banji')
        if request.POST['inlineRadioOptions']=='option1':
            stu_name1 = request.POST.get('name1')
            stu_name2 = request.POST.get('name2')
            stu_name3 = request.POST.get('name3')
            score1 = request.POST.get('score1')
            score2 = request.POST.get('score2')
            score3 = request.POST.get('score3')
        else:
            stu_name1 = request.POST.get('stu1')
            stu_name2 = request.POST.get('stu2')
            stu_name3 = request.POST.get('stu3')
            score1 = request.POST.get('sco1')
            score2 = request.POST.get('sco2')
            score3 = request.POST.get('sco3')

        if act_name and banji_name and stu_name1 and stu_name2 and stu_name3 and score1 and score2 and score3:
            banji = Banji.objects.get(name=banji_name)
            act = Act()
            act.banji = banji
            act.title = act_name
            act.save()

            student1 = get_object_or_404(Student, name=stu_name1)
            student2 = get_object_or_404(Student, name=stu_name2)
            student3 = get_object_or_404(Student, name=stu_name3)


            ss1 = StudentActScore()
            ss1.student = student1
            ss1.act = act
            ss1.score = score1
            ss1.save()

            ss2 = StudentActScore()
            ss2.student = student2
            ss2.act = act
            ss2.score = score2
            ss2.save()

            ss3 = StudentActScore()
            ss3.student = student3
            ss3.act = act
            ss3.score = score3
            ss3.save()
            return JsonResponse({'status': 'success'})
        else:
            return JsonResponse({'status': 'fail'})


    # login_url = '/login/'
    def get(self, request):
        courses = Schedule.objects.filter(teacher=request.user)
        ones = courses.filter(order='一')
        twos = courses.filter(order='二')
        threes = courses.filter(order='三')
        fours = courses.filter(order='四')
        fives = courses.filter(order='五')
        sixes = courses.filter(order='六')

        weekday = time.strftime("%w")
        weekdict = {'0': '星期日',
                    '1': '星期一',
                    '2': '星期二',
                    '3': '星期三',
                    '4': '星期四',
                    '5': '星期五',
                    '6': '星期六'
                    }
        today = weekdict[weekday]
        today_courses = courses.filter(week=today)
        today_courses = self.modify_day_courses(today_courses)
        while not any(today_courses):
            weekday = str((int(weekday) + 1) % 7)
            weekdict = {'0': '星期日',
                        '1': '星期一',
                        '2': '星期二',
                        '3': '星期三',
                        '4': '星期四',
                        '5': '星期五',
                        '6': '星期六'
                        }
            today = weekdict[weekday]
            today_courses = courses.filter(week=today)
            today_courses = self.modify_day_courses(today_courses)

        before_course=None
        current=time.strftime('%H')
        if int(current)< 10:
            before_course=today_courses[0]
        elif int(current) < 12:
            part_courses = today_courses[0:2]
            before_course = self.first_no_none(part_courses)
        elif int(current) < 15:
            part_courses = today_courses[0:3]
            before_course = self.first_no_none(part_courses)
        elif int(current) < 17:
            part_courses = today_courses[0:4]
            before_course = self.first_no_none(part_courses)
        elif int(current) < 19:
            part_courses = today_courses[0:5]
            before_course = self.first_no_none(part_courses)
        elif int(current) < 24:
            part_courses = today_courses[:]
            before_course = self.first_no_none(part_courses)

        # 获取当前上课班级的学生名单
        students = Student.objects.filter(banji=before_course.banji)[:]
        # 随机返回
        new_students=[s for s in students]

        students = random.sample(new_students, 3)
        return render(request, 'teacher/classactive.html', {'before_course': before_course,
                                                            'students': students})

    def first_no_none(self, l):
        t = l[::-1]
        for x in t:
            if x:
                break
        return x

    def modify_day_courses(self, ones):
        all_day = [None, None, None, None, None, None]
        for c in ones:
            if c.order == '一':
                all_day[0] = c
            elif c.order == '二':
                all_day[1] = c
            elif c.order == '三':
                all_day[2] = c
            elif c.order == '四':
                all_day[3] = c
            elif c.order == '五':
                all_day[4] = c
            else:
                all_day[5] = c

        return all_day


class CourseTableView(LoginRequiredMixin, View):
    # login_url = '/login/'
    def get(self, request):
        courses = Schedule.objects.filter(teacher=request.user)
        ones = courses.filter(order='一')
        twos = courses.filter(order='二')
        threes = courses.filter(order='三')
        fours = courses.filter(order='四')
        fives = courses.filter(order='五')
        sixes = courses.filter(order='六')

        weekday = time.strftime("%w")
        weekdict = {'0': '星期日',
                    '1': '星期一',
                    '2': '星期二',
                    '3': '星期三',
                    '4': '星期四',
                    '5': '星期五',
                    '6': '星期六'
                    }
        today = weekdict[weekday]
        today_courses = courses.filter(week=today)
        today_courses = self.modify_day_courses(today_courses)

        ones = self.modify_courses(ones)
        twos = self.modify_courses(twos)
        threes = self.modify_courses(threes)
        fours = self.modify_courses(fours)
        fives = self.modify_courses(fives)
        sixes = self.modify_courses(sixes)
        return render(request, 'teacher/coursetable.html', {'ones': ones,
                                                            'twos': twos,
                                                            'threes': threes,
                                                            'fours': fours,
                                                            'fives': fives,
                                                            'sixes': sixes,
                                                            'today_courses': today_courses,
                                                            })

    def modify_courses(self, ones):
        all_ones = [None, None, None, None, None, None, None]
        for c in ones:
            if c.week == '星期一':
                all_ones[0] = c
            elif c.week == '星期二':
                all_ones[1] = c
            elif c.week == '星期三':
                all_ones[2] = c
            elif c.week == '星期四':
                all_ones[3] = c
            elif c.week == '星期五':
                all_ones[4] = c
            elif c.week == '星期六':
                all_ones[5] = c
            else:
                all_ones[6] = c
        return all_ones

    def modify_day_courses(self, ones):
        all_day = [None, None, None, None, None, None]
        for c in ones:
            if c.order == '一':
                all_day[0] = c
            elif c.order == '二':
                all_day[1] = c
            elif c.order == '三':
                all_day[2] = c
            elif c.order == '四':
                all_day[3] = c
            elif c.order == '五':
                all_day[4] = c
            else:
                all_day[5] = c

        return all_day


class LogoutView(View):
    def post(self, request):
        logout(request)
        return redirect('teacher:index')


class CustomerAskView(View):
    def post(self, request):
        customerask_form = CustomerForm(request.POST)
        if customerask_form.is_valid():
            customerask_form.save()
            return JsonResponse({'status': 'success'})
        else:
            return JsonResponse({'status': 'fail', 'msg': '提交出错了'})


class SummaryView(LoginRequiredMixin, View):
    def get(self, request):

        t_schedules = Schedule.objects.filter(teacher=request.user)
        t_schedules = t_schedules.values('banji__name', 'course__name').distinct()

        return render(request, 'teacher/summaryscore.html', {'t_schedules': t_schedules})


class ClassStudentActScoreView(View):
    def get(self, request, banji_name, scorename):
        banji = get_object_or_404(Banji, name=banji_name)
        scorename=scorename
        students = banji.student_set.all()
        acts = banji.act_set.all()
        acts = acts.filter(teacher=request.user)

        stuactscore=[]
        for stu in students:
            temp = []
            temp.append(stu.No)
            temp.append(stu.name)
            for act in acts:
                try:
                    stuact = StudentActScore.objects.get(student=stu, act=act)
                    temp.append(stuact.score)
                except StudentActScore.DoesNotExist:
                    temp.append(0)
            stuactscore.append(temp)

        return render(request, 'teacher/classscore.html', {'students': students, 'acts': acts,
                                                           'stuactscore': stuactscore,
                                                           'banji': banji,
                                                           'coursename': scorename,
                                                           'banji_name': banji_name,
                                                           'scorename': scorename})


def export_to_excel(request, banji_name, scorename):
    banji = get_object_or_404(Banji, name=banji_name)
    scorename = scorename
    students = banji.student_set.all()
    acts = banji.act_set.all()
    acts = acts.filter(teacher=request.user)

    stuactscore = []
    for stu in students:
        temp = []
        temp.append(stu.No)
        temp.append(stu.name)
        for act in acts:
            try:
                stuact = StudentActScore.objects.get(student=stu, act=act)
                temp.append(stuact.score)
            except StudentActScore.DoesNotExist:
                temp.append(0)
        stuactscore.append(temp)

    wb = Workbook()
    ws = wb.active
    ws.title = 'New Title'
    ws.cell(row=2, column=1).value = '学号'
    ws.cell(row=2, column=2).value = '姓名'
    i = 1
    for act in acts:
        ws.cell(row=2, column=2+i).value = act.title + '(' + act.created.strftime('%d/%m') + ')'
        i = i + 1
    m = 0

    for ss in stuactscore:
        m = m + 1
        n = 0
        for d in ss:
            n = n + 1
            ws.cell(row=2+m, column=n).value = d
    max_col = ws.max_column
    max_row = ws.max_row

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
    ws.cell(row=1, column=1).value = scorename + "平时成绩"
    ws.cell(row=1, column=1).font = Font(size=20)
    ws.cell(row=1, column=1).alignment = Alignment(horizontal='center', vertical='center')

    # ws.cell(row=max_row+2, column=max_col-4).value = '班级：' + banji.name
    # ws.cell(row=max_row + 3, column=max_col - 4).value = '日期：' + time.strftime("%d/%m")

    response = HttpResponse(content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filename="test.xls"'
    wb.save(response)
    return response